import json
import sys
import time
from datetime import date
from urllib.request import urlopen

import openpyxl
from common import globalConfig

def getCodeStr(code):
    head = code[0]
    if head == "0" or head == "3":
        return "SZ{}".format(code)
    else:
        return "SH{}".format(code)

def getReport(code):
    codeStr = getCodeStr(code)
    url = "http://f10.eastmoney.com/ProfitForecast/ProfitForecastAjax?code={}".format(codeStr)
    response = urlopen(url)
    # print(response.read())
    a = json.loads(response.read())
    if "pjtj" in a:
        return a['pjtj']
    else:
        return None

def setPJTJInExl(code, wb:openpyxl.Workbook):
    pjtj = getReport(code)
    if pjtj == None:
        return
    i = 0
    for ele in pjtj:
        wb.worksheets[i].append((str(code), float(ele["pjxs"]), int(ele["mr"]), int(ele["zc"]), int(ele["zx"]),int(ele["jc"]),int(ele["mc"]),int(ele["zjs"])))
        i = i + 1


def makeExl():
    wb = openpyxl.Workbook()
    wb.active.title = "1月"
    wb.active.append(("code", "pjxs", "mr", "zc", "zx", "jc", "mc", "zjs"))
    wb.create_sheet("2月").append(("code", "pjxs", "mr", "zc", "zx", "jc", "mc", "zjs"))
    wb.create_sheet("3月").append(("code", "pjxs", "mr", "zc", "zx", "jc", "mc", "zjs"))
    wb.create_sheet("4月").append(("code", "pjxs", "mr", "zc", "zx", "jc", "mc", "zjs"))
    wb.create_sheet("5月").append(("code", "pjxs", "mr", "zc", "zx", "jc", "mc", "zjs"))
    wb.create_sheet("6月").append(("code", "pjxs", "mr", "zc", "zx", "jc", "mc", "zjs"))
    wb.create_sheet("1年").append(("code", "pjxs", "mr", "zc", "zx", "jc", "mc", "zjs"))
    return wb

def main():
    wb = makeExl()
    for head in globalConfig.heads:
        for i in range(1000):
            code = "{}{}".format(head, str(i).zfill(3))
            if not code in globalConfig.blackList:
                setPJTJInExl(code, wb)
    wb.save("report/report_{}.xlsx".format(date.today()))

print("start! {}".format(time.strftime("%Y/%m/%d %H:%M:%S")))
main()
print("over! {}".format(time.strftime("%Y/%m/%d %H:%M:%S")))
sys.exit()